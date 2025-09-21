from django.shortcuts import render
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from datetime import datetime
import os 
from openpyxl import Workbook, load_workbook # Import the necessary classes

# Create your views here.
def home_view(request):
    return render(request, 'index.html')

@csrf_exempt
def contact_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            name = data.get('name')
            email = data.get('email')
            message = data.get('message')

            if not all([name, email, message]):
                return JsonResponse({'status': 'error', 'message': 'All fields are required.'}, status=400)

            # --- New Code to Save to an Excel Document (.xlsx) ---
            submissions_dir = os.path.join(settings.BASE_DIR, 'myapp', 'submissions')
            if not os.path.exists(submissions_dir):
                os.makedirs(submissions_dir)

            filename = 'contact_submissions.xlsx'
            filepath = os.path.join(submissions_dir, filename)

            # Check if the Excel file already exists
            if os.path.exists(filepath):
                # If it exists, load the workbook
                workbook = load_workbook(filepath)
                sheet = workbook.active
            else:
                # If it doesn't exist, create a new workbook and add headers
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(['Timestamp', 'Name', 'Email', 'Message'])

            # Append the new data to the next empty row
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sheet.append([timestamp, name, email, message])

            # Save the workbook
            workbook.save(filepath)
            # --- End of New Code ---

            return JsonResponse({'status': 'success', 'message': 'Message saved successfully to Excel!'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)